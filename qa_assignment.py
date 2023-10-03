import openpyxl
from selenium import webdriver
import datetime
import os

def get_current_date():
  return datetime.date.today()

def file_exists(file_path):
  return os.path.isfile(file_path)

def read_keywords_and_dates_from_excel(excel_file_path):
  wb = openpyxl.Workbook(excel_file_path)
  sheet = wb.active

  keywords = []
  dates = []
  for row in sheet.rows:
    keyword = row[0].value
    date = row[1].value

    keywords.append(keyword)
    dates.append(date)

  return keywords, dates

def search_for_keyword_on_google(keyword):
  driver = webdriver.Chrome()
  driver.get("https://www.google.com/")

  search_box = driver.find_element_by_name("q")
  search_box.send_keys(keyword)
  search_box.submit()

  return driver

def find_longest_and_shortest_options_from_search_results(driver):
  longest_option = driver.find_element_by_css_selector(".r a:nth-child(1)")
  shortest_option = driver.find_element_by_css_selector(".r a:last-child")

  return longest_option.text, shortest_option.text

def write_longest_and_shortest_options_to_excel(longest_option, shortest_option, excel_file_path, row_index, column_index):
  wb = openpyxl.Workbook(excel_file_path)
  sheet = wb.active

  sheet.cell(row=row_index, column=column_index).value = longest_option
  sheet.cell(row=row_index, column=column_index + 1).value = shortest_option

  wb.save(excel_file_path)

def main():
  # Check if the Excel file exists.
  if not file_exists("keywords.xlsx"):
    print("The Excel file 'keywords.xlsx' does not exist.")
    return

  # Read the list of keywords and their corresponding dates from the Excel file.
  keywords, dates = read_keywords_and_dates_from_excel("keywords.xlsx")

  # For each keyword and date, search for the keyword on Google, find the longest and shortest options from the search results, and write the longest and shortest options to the Excel file on the corresponding row and column.
  for keyword, date in zip(keywords, dates):
    driver = search_for_keyword_on_google(keyword)

    longest_option, shortest_option = find_longest_and_shortest_options_from_search_results(driver)

    write_longest_and_shortest_options_to_excel(longest_option, shortest_option, "keywords.xlsx", date.row, 2)

if __name__ == "__main__":
  main()
